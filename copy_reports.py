#!/usr/bin/env python3
"""
Script to copy PDF files from ZZZmemories folders to corresponding student directories.
"""

import os
import shutil
import glob
from pathlib import Path
from openpyxl import load_workbook

def get_students_from_committees():
    """Read student names from committees.xlsx file."""
    try:
        # Load the workbook and get the active sheet
        wb = load_workbook('committees.xlsx')
        ws = wb.active
        
        # Find the "Author" column (assuming it's the first column with "Author" header)
        author_col = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == 'Author':
                author_col = col
                break
        
        if author_col is None:
            print("❌ 'Author' column not found in committees.xlsx")
            return []
        
        # Read all student names from the Author column (skip header row)
        students = []
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=author_col).value
            if cell_value and str(cell_value).strip():
                students.append(str(cell_value).strip())

        return students
    except Exception as e:
        print(f"❌ Error reading committees.xlsx: {e}")
        return []

def normalize_name(name):
    """Normalize student names for matching."""
    # Remove accents and special characters for comparison
    replacements = {
        'á': 'a', 'à': 'a', 'ä': 'a', 'â': 'a',
        'é': 'e', 'è': 'e', 'ë': 'e', 'ê': 'e',
        'í': 'i', 'ì': 'i', 'ï': 'i', 'î': 'i',
        'ó': 'o', 'ò': 'o', 'ö': 'o', 'ô': 'o',
        'ú': 'u', 'ù': 'u', 'ü': 'u', 'û': 'u',
        'ñ': 'n', 'ç': 'c'
    }
    
    normalized = name.lower()
    for old, new in replacements.items():
        normalized = normalized.replace(old, new)
    
    return normalized

def extract_student_name_from_folder(folder_name):
    """Extract student name from ZZZmemories folder format."""
    # Format: "Last Name, First Name_ID_assignsubmission_file"
    if '_' in folder_name:
        name_part = folder_name.split('_')[0]  # Get everything before first underscore
        if ',' in name_part:
            last, first = name_part.split(',', 1)
            return f"{first.strip()}_{last.strip()}"
    return folder_name

def find_matching_student_dir(student_name, student_dirs, committee_students):
    """Find matching student directory."""
    normalized_name = normalize_name(student_name)
    
    # First, check if this student is in the committees list
    student_in_committee = False
    committee_student_match = None
    
    for committee_student in committee_students:
        normalized_committee = normalize_name(committee_student)
        # Check if the extracted name matches a committee student
        if normalized_name.replace('_', ' ') in normalized_committee or normalized_committee in normalized_name.replace('_', ' '):
            student_in_committee = True
            committee_student_match = committee_student
            break
    
    if not student_in_committee:
        return None, f"Student not found in committees.xlsx"
    
    for student_dir in student_dirs:
        normalized_dir = normalize_name(student_dir)
        
        # Try exact match first
        if normalized_name == normalized_dir:
            return student_dir, f"Matched with committee student: {committee_student_match}"
            
        # Try partial matches
        name_parts = normalized_name.replace('_', ' ').split()
        dir_parts = normalized_dir.replace('_', ' ').split()
        
        # Check if all name parts are in directory name
        if all(part in ' '.join(dir_parts) for part in name_parts):
            return student_dir, f"Matched with committee student: {committee_student_match}"
    
    return None, f"Directory not found for committee student: {committee_student_match}"

def copy_pdfs():
    """Main function to copy PDF files."""
    script_dir = Path(__file__).parent
    memories_dir = script_dir / "ZZZmemories"
    
    if not memories_dir.exists():
        print(f"❌ ZZZmemories directory not found: {memories_dir}")
        return
    
    # Read students from committees.xlsx
    committee_students = get_students_from_committees()
    if not committee_students:
        print("❌ No students found in committees.xlsx")
        return
    
    # Get all student directories (exclude special directories)
    exclude_dirs = {'.git', '__pycache__', 'ZZZmemories'}
    student_dirs = [d.name for d in script_dir.iterdir() 
                   if d.is_dir() and d.name not in exclude_dirs and not d.name.startswith('.')]
    
    print(f"\nFound {len(student_dirs)} student directories")
    
    # Get all memory folders
    memory_folders = [d for d in memories_dir.iterdir() if d.is_dir()]
    print(f"Found {len(memory_folders)} memory folders")
    
    copied_count = 0
    not_found_count = 0
    
    # Lists to track results
    successfully_copied = []
    not_matched = []
    
    for memory_folder in memory_folders:
        print(f"📁 Processing: {memory_folder.name}")
        
        # Extract student name from folder
        student_name = extract_student_name_from_folder(memory_folder.name)
        
        # Find matching student directory
        matching_dir, match_reason = find_matching_student_dir(student_name, student_dirs, committee_students)
        
        if not matching_dir:
            print(f"   ❌ {match_reason}")
            not_found_count += 1
            not_matched.append({
                'memory_folder': memory_folder.name,
                'extracted_name': student_name,
                'reason': match_reason,
                'available_dirs': student_dirs
            })
            continue
        
        # Find PDF files in memory folder
        pdf_files = list(memory_folder.glob("*.pdf"))
        
        if not pdf_files:
            print(f"   ⚠️  No PDF files found in {memory_folder.name}")
            continue
        
        # Copy PDF files to student directory
        target_dir = script_dir / matching_dir
        
        for pdf_file in pdf_files:
            target_path = target_dir / pdf_file.name
            
            try:
                shutil.copy2(pdf_file, target_path)
                #print(f"   📄 Copied: {pdf_file.name} → {matching_dir}/")
                copied_count += 1
            except Exception as e:
                print(f"   ❌ Error copying {pdf_file.name}: {e}")
        
        # Track successful copy
        successfully_copied.append({
            'memory_folder': memory_folder.name,
            'extracted_name': student_name,
            'target_directory': matching_dir,
            'pdf_files': [f.name for f in pdf_files]
        })
    
    print(f"\n🎉 Summary:")
    print(f"   ✅ Successfully copied: {copied_count} PDF files")
    print(f"   ❌ No matching directory: {not_found_count} students")
    
    # Detailed summary
    
    if not_matched:
        print(f"❌ MEMORIES NOT COPIED ({len(not_matched)} students):")
        print("-" * 60)
        for i, item in enumerate(not_matched, 1):
            print(f"{i:2d}. {item['extracted_name']}")
            print(f"    Memory folder: {item['memory_folder']}")
            print(f"    Reason: {item['reason']}")
            if 'not found in committees.xlsx' not in item['reason']:
                print(f"    Available directories that might match:")
                # Try to find potential matches
                potential_matches = []
                normalized_name = normalize_name(item['extracted_name'])
                for dir_name in item['available_dirs']:
                    normalized_dir = normalize_name(dir_name)
                    if any(part in normalized_dir for part in normalized_name.replace('_', ' ').split()):
                        potential_matches.append(dir_name)
                
                if potential_matches:
                    for match in potential_matches[:3]:  # Show max 3 potential matches
                        print(f"      - {match}")
                else:
                    print(f"      - No similar directories found")
            print()
    else:
        print(f"🎉 ALL MEMORIES SUCCESSFULLY COPIED!")
    
    print("="*80)

if __name__ == "__main__":
    copy_pdfs()
