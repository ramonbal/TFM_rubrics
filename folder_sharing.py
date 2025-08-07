#!/usr/bin/env python3
"""
OneDrive Folder Sharing Automation
Automates the process of sharing OneDrive folders with configurable permissions.

Requirements:
- pip install pyperclip openpyxl pyautogui

Features:
- Automated folder sharing with OneDrive
- Configurable sharing permissions
- Excel export of results
- Error handling and logging
"""

import os
import time
import pyperclip
import pyautogui
from pathlib import Path
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Configure pyautogui for safety and stability
pyautogui.FAILSAFE = True  # Move mouse to top-left corner to stop
pyautogui.PAUSE = 1.0

class FolderSharer:
    def __init__(self):
        self.sharing_links = {}
        self.current_folder_idx = 0
        self.folders = []
        
    def check_subfolders_inside(self, folder_path):
        """Check if a folder contains subfolders and return their names."""
        subfolders = []
        folder_path = Path(folder_path)
        
        for item in folder_path.iterdir():
            if item.is_dir() and not item.name.startswith('.'):
                subfolders.append(item)
        
        return subfolders
    
    def get_subfolders(self, base_path):
        """Get all subfolders to process and check for nested subfolders."""
        base_path = Path(base_path)
        folders = {}
        
        for item in base_path.iterdir():
            if (item.is_dir() and 
                not item.name.startswith('.') and 
                item.name not in ['__pycache__', 'ZZZmemories']):
                
                # Check for subfolders inside this folder
                nested_subfolders = self.check_subfolders_inside(item)
                
                # Add to folders dict
                folders[item.name] = [item] + nested_subfolders
        
        return folders
    
    def open_folder_location(self, folder_path):
        """Open the folder location in Explorer."""
        try:
            os.system(f'explorer /select,"{folder_path}"')
            time.sleep(0.5)  # Wait for Explorer to open
            return True
        except Exception as e:
            print(f"❌ Error opening folder: {e}")
            return False
    
    def get_share_link(self, folder_path, position=16):
        """Automated sharing process using keyboard navigation."""
        folder_name = folder_path.name
        
        try:
            # Clear clipboard
            pyperclip.copy('')
            
            # Open context menu (Shift+F10)
            pyautogui.hotkey('shift', 'f10')
            time.sleep(1)
            
            # Navigate to Share option (typically 16 downs in OneDrive folders)
            pyautogui.press('down', presses=position)
            pyautogui.press('enter')
            pyautogui.press(['down', 'up'])
            pyautogui.press('enter')
            time.sleep(2)

            print(f"Active window title: {pyautogui.getActiveWindow().title}")
            if not (folder_path.name.lower() in pyautogui.getActiveWindow().title.lower()):
                pyautogui.hotkey('alt', 'f4')
                return self.get_share_link(folder_path, position=15)

            # Set permissions to "Can edit"
            pyautogui.press('tab', presses=4)
            pyautogui.press('enter')
            pyautogui.press('up', presses=3)
            pyautogui.press(['enter'])
            #qualsevol que tingui el link
            pyautogui.press('tab', presses=4)
            pyautogui.press('enter')
            pyautogui.press('up', presses=4)
            pyautogui.press('tab', presses=5)
            pyautogui.press('enter', presses=3)
            time.sleep(0.2)
            
            # Close sharing dialog
            pyautogui.hotkey('alt', 'f4')
            
            # Verify link was copied
            clipboard_content = pyperclip.paste()
            if (clipboard_content and 
                ('sharepoint.com' in clipboard_content or 'onedrive.live.com' in clipboard_content) and clipboard_content):
                
                print(f"✅ Successfully shared: {folder_name}")
                return clipboard_content
            else:
                print(f"⚠️  Failed to get sharing link for: {folder_name}")        
                return "NOT PROCESSED"
                
        except Exception as e:
            print(f"❌ Error sharing folder {folder_name}: {e}")
            return "NOT PROCESSED"

    def process_folder(self, folder_name):
        """Process a single folder."""
        
        self.sharing_links[folder_name] = []
        for folder_path in self.folders[folder_name]:
            # Open folder location
            print(f"📁 Opening folder: {folder_path}")
            if not self.open_folder_location(folder_path):
                print(f"❌ Failed to open folder location for: {folder_name}")
                return False
            
            # Attempt to share the folder
            success = self.get_share_link(folder_path)
            
            self.sharing_links[folder_name] += [success]
        
            # Close Explorer window
#            pyautogui.hotkey('alt', 'f4')
        
        return True
    
    def save_results(self):
        """Save results to Excel file."""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Folder_Sharing_Results"
            
            # Headers
            headers = ["Folder Name", "Sharing Link"]
            ws.append(headers)
            
            for folder_name in self.sharing_links:
                link = self.sharing_links[folder_name]
                ws.append([folder_name]+ link)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 100)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            filename = f"folder_sharing_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            
            print(f"\n💾 Results saved to: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ Error saving results: {e}")
            return None
    
    def load_folder_links_from_excel(self, excel_file=None):
        """Load names of failed and successful folders from Excel file."""
        try:
            import glob
            from openpyxl import load_workbook
            
            # If no specific file provided, find the most recent one
            if excel_file is None:
                excel_files = glob.glob("folder_sharing_*.xlsx")
                if not excel_files:
                    print("❌ No Excel results files found")
                    return False
                
                excel_file = max(excel_files, key=os.path.getctime)
                print(f"📁 Using Excel file: {excel_file}")
            
            # Load the Excel file
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # Skip header row, start from row 2
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 2:  # Make sure we have enough columns
                    folder_name = row[0]  # First column is folder name
                    sharing_link = row[1]  # Second column is sharing link
                    # Check if this folder failed (no valid sharing link)
                    if not (sharing_link and sharing_link.startswith('http')):
                        sharing_link = None                     
                    self.sharing_links[folder_name] = [sharing_link]
                    sharing_link = row[2]  # Second column is sharing link
                    # Check if this folder failed (no valid sharing link)
                    if sharing_link and sharing_link.startswith('http'):
                        self.sharing_links[folder_name] += [sharing_link]
                    
            return True
            
        except Exception as e:
            print(f"❌ Error loading Excel file: {e}")
            return False

    def run(self, base_path, max_folders=None, retry_folder_names=False):
        """Run the folder sharing process."""
        # Get folders to process
        self.folders = self.get_subfolders(base_path)
        if len(self.sharing_links) > 0:
            self.folders = {name: path for name, path in self.folders.items() if name in self.sharing_links and self.sharing_links[name][0] is None}
                    
        if not self.folders:
            print("❌ No folders found to process")
            return
        
        print(f"\n📁 Found {len(self.folders)} folders to share:")
        for i, folder_name in enumerate(self.folders.keys(), 1):
            print(f"   {i:2d}. {folder_name} {'advisor' if len(self.folders[folder_name]) > 1 else ''}")

        print(f"\n⚙️  This script will:")
        print(f"   • Open each folder in Explorer")
        print(f"   • Configure sharing as 'Anyone with the link, Can edit'")
        print(f"   • Copy sharing links automatically")
        print(f"   • Save results to Excel")
        
        print(f"\n⚠️  Important:")
        print(f"   • Don't move mouse or type during automation")
        print(f"   • Move mouse to top-left corner to emergency stop")
        
#        response = input(f"\nStart sharing process? (y/n): ")
#        if not response.lower().startswith('y'):
#            print("Process cancelled.")
#            return
        
        # Process each folder - compact loop
        for i, folder_name in enumerate(self.folders):
            if max_folders is not None and i >= max_folders:
                break
            self.current_folder_idx = i
            self.process_folder(folder_name)
            
            # Small delay between folders
            if i < len(self.folders) - 1:
                time.sleep(0.2)
        
        # Save final results
        self.save_results()


def main():
    """Main function."""
    current_path = Path.cwd()
    print(f"📁 Working directory: {current_path}")
    
    # Create and run sharer
    sharer = FolderSharer()
    
    failed_names = sharer.load_folder_links_from_excel()
    sharer.run(current_path, max_folders=4, retry_folder_names=failed_names)


if __name__ == "__main__":
    main()
